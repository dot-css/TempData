"""
Excel export functionality

Exports data to Excel format with proper formatting and data type handling.
"""

import pandas as pd
import os
from typing import Optional, Union, Dict, Any, List
from datetime import datetime
from .base_exporter import BaseExporter


class ExcelExporter(BaseExporter):
    """
    Exporter for Excel format
    
    Handles Excel export with proper formatting, data types, and worksheet management.
    Supports multiple sheets, styling, and Excel-specific features.
    """
    
    def __init__(self):
        """Initialize Excel exporter"""
        super().__init__()
        self.supported_extensions = ['.xlsx', '.xls']
        
        # Default Excel export options
        self.default_options = {
            'index': False,
            'header': True,
            'sheet_name': 'Sheet1',
            'engine': 'openpyxl',  # Use openpyxl for .xlsx files
            'na_rep': '',  # How to represent NaN values
        }
        
        # Default styling options
        self.default_styling = {
            'header_style': {
                'bold': True,
                'bg_color': '#D3D3D3',
                'font_color': '#000000',
                'border': 1
            },
            'auto_adjust_width': True,
            'freeze_panes': (1, 0),  # Freeze header row
            'add_filters': True
        }
    
    def export(self, data: pd.DataFrame, filename: str, **kwargs) -> str:
        """
        Export data to Excel format with proper formatting
        
        Args:
            data: DataFrame to export
            filename: Output filename
            **kwargs: Excel-specific options that override defaults
                - sheet_name: Name of the worksheet (default: 'Sheet1')
                - index: Include row index (default: False)
                - header: Include column headers (default: True)
                - engine: Excel engine to use ('openpyxl' or 'xlsxwriter')
                - date_format: Date formatting string
                - datetime_format: Datetime formatting string
                - na_rep: String representation of NaN values
                - apply_styling: Apply default styling (default: True)
                - auto_adjust_width: Auto-adjust column widths (default: True)
                - freeze_panes: Freeze panes tuple (row, col) (default: (1, 0))
                - add_filters: Add autofilters to headers (default: True)
                - header_style: Dictionary of header styling options
                
        Returns:
            str: Path to exported Excel file
            
        Raises:
            ValueError: If data is invalid or filename is invalid
            IOError: If file cannot be written
            ImportError: If required Excel libraries are not installed
        """
        # Check for required dependencies
        self._check_excel_dependencies()
        
        # Validate input data
        self._validate_data(data)
        
        # Validate and normalize filename
        if not filename.endswith('.xlsx') and not filename.endswith('.xls'):
            excel_path = self._validate_filename(filename, '.xlsx')
        else:
            excel_path = self._validate_filename(filename)
        
        # Ensure output directory exists
        self._ensure_directory_exists(excel_path)
        
        # Prepare data for Excel export
        export_data = self._prepare_excel_data(data)
        
        # Merge default options with user-provided options
        export_options = {**self.default_options, **kwargs}
        apply_styling = export_options.pop('apply_styling', True)
        
        # Extract styling options
        styling_options = {**self.default_styling}
        for key in ['auto_adjust_width', 'freeze_panes', 'add_filters', 'header_style']:
            if key in kwargs:
                styling_options[key] = kwargs.pop(key)
        
        try:
            # Export to Excel with proper error handling
            if apply_styling and export_options.get('engine') == 'openpyxl':
                # Use openpyxl for advanced styling
                self._export_with_styling(export_data, excel_path, export_options, styling_options)
            else:
                # Basic export without styling
                export_data.to_excel(excel_path, **export_options)
            
            # Verify the file was created successfully
            if not os.path.exists(excel_path):
                raise IOError(f"Failed to create Excel file: {excel_path}")
                
            return excel_path
            
        except Exception as e:
            # Clean up partial file if it exists
            if os.path.exists(excel_path):
                try:
                    os.remove(excel_path)
                except OSError:
                    pass
            raise IOError(f"Failed to export Excel file: {str(e)}")
    
    def _check_excel_dependencies(self):
        """
        Check if required Excel libraries are installed
        
        Raises:
            ImportError: If required libraries are missing
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install it with: pip install openpyxl"
            )
    
    def _prepare_excel_data(self, data: pd.DataFrame) -> pd.DataFrame:
        """
        Prepare data for Excel export with proper formatting
        
        Args:
            data: Original DataFrame
            
        Returns:
            pd.DataFrame: Prepared DataFrame for Excel export
        """
        # Create a copy to avoid modifying original data
        excel_data = data.copy()
        
        # Handle different data types appropriately for Excel
        for col in excel_data.columns:
            col_data = excel_data[col]
            
            # Handle datetime columns - Excel handles these well natively
            if pd.api.types.is_datetime64_any_dtype(col_data):
                # Ensure timezone-naive datetimes for Excel compatibility
                if hasattr(col_data.dtype, 'tz') and col_data.dtype.tz is not None:
                    excel_data[col] = col_data.dt.tz_convert(None)
            
            # Handle boolean columns - keep as boolean
            elif pd.api.types.is_bool_dtype(col_data):
                pass
            
            # Handle numeric columns
            elif pd.api.types.is_numeric_dtype(col_data):
                # Excel handles NaN well, no special processing needed
                pass
            
            # Handle object/string columns
            elif col_data.dtype == 'object':
                # Handle None/NaN values
                excel_data[col] = col_data.fillna('')
                
                # Convert to string to ensure consistency
                excel_data[col] = excel_data[col].astype(str)
                
                # Handle very long strings (Excel has a 32,767 character limit per cell)
                excel_data[col] = excel_data[col].apply(
                    lambda x: x[:32767] if len(str(x)) > 32767 else x
                )
        
        return excel_data
    
    def _export_with_styling(self, data: pd.DataFrame, excel_path: str, 
                           export_options: Dict[str, Any], styling_options: Dict[str, Any]):
        """
        Export Excel file with advanced styling using openpyxl
        
        Args:
            data: Prepared DataFrame
            excel_path: Output file path
            export_options: Export options
            styling_options: Styling options
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = export_options.get('sheet_name', 'Sheet1')
        
        # Add data to worksheet
        include_header = export_options.get('header', True)
        include_index = export_options.get('index', False)
        
        for r in dataframe_to_rows(data, index=include_index, header=include_header):
            ws.append(r)
        
        # Apply styling
        if include_header:
            self._apply_header_styling(ws, styling_options.get('header_style', {}))
        
        if styling_options.get('auto_adjust_width', True):
            self._auto_adjust_column_widths(ws)
        
        if styling_options.get('freeze_panes'):
            freeze_row, freeze_col = styling_options['freeze_panes']
            ws.freeze_panes = ws.cell(row=freeze_row + 1, column=freeze_col + 1)
        
        if styling_options.get('add_filters', True) and include_header:
            ws.auto_filter.ref = ws.dimensions
        
        # Save workbook
        wb.save(excel_path)
    
    def _apply_header_styling(self, worksheet, header_style: Dict[str, Any]):
        """
        Apply styling to header row
        
        Args:
            worksheet: openpyxl worksheet
            header_style: Dictionary of styling options
        """
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        
        # Default header styling
        font = Font(
            bold=header_style.get('bold', True),
            color=header_style.get('font_color', '000000').replace('#', '')
        )
        
        fill = PatternFill(
            start_color=header_style.get('bg_color', 'D3D3D3').replace('#', ''),
            end_color=header_style.get('bg_color', 'D3D3D3').replace('#', ''),
            fill_type='solid'
        )
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        ) if header_style.get('border', True) else None
        
        alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply to first row (header)
        for cell in worksheet[1]:
            cell.font = font
            cell.fill = fill
            if border:
                cell.border = border
            cell.alignment = alignment
    
    def _auto_adjust_column_widths(self, worksheet):
        """
        Auto-adjust column widths based on content
        
        Args:
            worksheet: openpyxl worksheet
        """
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set width with some padding, but cap at reasonable maximum
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def export_multiple_sheets(self, data_dict: Dict[str, pd.DataFrame], 
                              filename: str, **kwargs) -> str:
        """
        Export multiple DataFrames to different sheets in one Excel file
        
        Args:
            data_dict: Dictionary mapping sheet names to DataFrames
            filename: Output filename
            **kwargs: Excel export options
            
        Returns:
            str: Path to exported Excel file
        """
        # Check dependencies
        self._check_excel_dependencies()
        
        # Validate filename
        excel_path = self._validate_filename(filename, '.xlsx')
        self._ensure_directory_exists(excel_path)
        
        # Validate all DataFrames
        for sheet_name, data in data_dict.items():
            self._validate_data(data)
        
        try:
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                for sheet_name, data in data_dict.items():
                    # Prepare data for each sheet
                    prepared_data = self._prepare_excel_data(data)
                    
                    # Export options for this sheet - only use pandas-compatible options
                    sheet_options = {
                        'index': kwargs.get('index', self.default_options['index']),
                        'header': kwargs.get('header', self.default_options['header']),
                        'na_rep': kwargs.get('na_rep', self.default_options['na_rep'])
                    }
                    
                    prepared_data.to_excel(writer, sheet_name=sheet_name, **sheet_options)
            
            return excel_path
            
        except Exception as e:
            if os.path.exists(excel_path):
                try:
                    os.remove(excel_path)
                except OSError:
                    pass
            raise IOError(f"Failed to export multi-sheet Excel file: {str(e)}")
    
    def export_with_charts(self, data: pd.DataFrame, filename: str, 
                          chart_configs: List[Dict[str, Any]], **kwargs) -> str:
        """
        Export Excel file with embedded charts
        
        Args:
            data: DataFrame to export
            filename: Output filename
            chart_configs: List of chart configuration dictionaries
            **kwargs: Excel export options
            
        Returns:
            str: Path to exported Excel file
        """
        # This is a placeholder for advanced charting functionality
        # Would require additional dependencies like xlsxwriter or advanced openpyxl usage
        # For now, export basic Excel file
        return self.export(data, filename, **kwargs)
    
    def get_export_info(self, data: pd.DataFrame) -> Dict[str, Any]:
        """
        Get information about what the Excel export will contain
        
        Args:
            data: DataFrame to analyze
            
        Returns:
            Dict with export information
        """
        info = {
            'rows': len(data),
            'columns': len(data.columns),
            'column_names': list(data.columns),
            'data_types': {},
            'memory_usage_mb': round(data.memory_usage(deep=True).sum() / (1024 * 1024), 2),
            'estimated_excel_size_mb': 0,
            'excel_limitations': {
                'max_rows': 1048576,
                'max_columns': 16384,
                'max_chars_per_cell': 32767
            }
        }
        
        # Analyze data types
        for col in data.columns:
            dtype = str(data[col].dtype)
            info['data_types'][col] = dtype
        
        # Check for Excel limitations
        if len(data) > info['excel_limitations']['max_rows']:
            info['warnings'] = info.get('warnings', [])
            info['warnings'].append(f"Data has {len(data)} rows, exceeding Excel's limit of {info['excel_limitations']['max_rows']}")
        
        if len(data.columns) > info['excel_limitations']['max_columns']:
            info['warnings'] = info.get('warnings', [])
            info['warnings'].append(f"Data has {len(data.columns)} columns, exceeding Excel's limit of {info['excel_limitations']['max_columns']}")
        
        # Estimate Excel file size (rough approximation)
        # Excel files are typically larger than CSV due to formatting and metadata
        avg_chars_per_cell = 12
        total_cells = len(data) * len(data.columns)
        estimated_size = total_cells * avg_chars_per_cell * 1.5  # Excel overhead
        info['estimated_excel_size_mb'] = round(estimated_size / (1024 * 1024), 2)
        
        return info